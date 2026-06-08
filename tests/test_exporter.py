import os, sys, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import openpyxl
from cashflow_tool.exporter import ExcelExporter
from cashflow_tool.models import CFRecord, MatchedPair, MatchResult

def test_export_creates_file():
    exporter = ExcelExporter()
    result = MatchResult(
        matched=[MatchedPair(pay_voucher='P1', payer='A公司', pay_cf='CF1', pay_amount=100,
                             recv_voucher='R1', receiver='B公司', recv_cf='CF2', recv_amount=-100,
                             match_type='精确匹配')],
        unmatched_pay=[CFRecord(company='C', counterparty='D', cf_full='CF3', amount=200,
                                 voucher_no='P3', direction='付款', source='测试')],
        unmatched_recv=[]
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    try:
        exporter.export(result, temp_path)
        assert os.path.exists(temp_path)
        assert os.path.getsize(temp_path) > 0
        wb = openpyxl.load_workbook(temp_path)
        assert '现金流内部抵消明细' in wb.sheetnames
        assert '未匹配明细' in wb.sheetnames
        assert '汇总表' in wb.sheetnames
        wb.close()
        print(f'OK: 文件大小 {os.path.getsize(temp_path)} bytes')
    finally:
        os.unlink(temp_path)


def test_export_empty_matched():
    """测试没有匹配记录的情况"""
    exporter = ExcelExporter()
    result = MatchResult(
        matched=[],
        unmatched_pay=[CFRecord(company='X', counterparty='Y', cf_full='CF99', amount=100,
                                 voucher_no='V1', direction='付款', source='测试')],
        unmatched_recv=[CFRecord(company='Y', counterparty='X', cf_full='CF88', amount=100,
                                  voucher_no='V2', direction='收款', source='测试')],
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    try:
        exporter.export(result, temp_path)
        wb = openpyxl.load_workbook(temp_path)
        assert '未匹配明细' in wb.sheetnames
        ws = wb['未匹配明细']
        # 检查未匹配明细有2行数据
        data_rows = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(row))
        assert data_rows == 2, f"期望2行未匹配数据，实际{data_rows}"
        wb.close()
    finally:
        os.unlink(temp_path)


def test_export_summary_values():
    """测试汇总表金额是否正确"""
    exporter = ExcelExporter()
    result = MatchResult(
        matched=[
            MatchedPair(pay_voucher='P1', payer='A', pay_cf='CF1', pay_amount=100,
                         recv_voucher='R1', receiver='B', recv_cf='CF2', recv_amount=-100,
                         match_type='精确匹配'),
            MatchedPair(pay_voucher='P2', payer='C', pay_cf='CF3', pay_amount=200,
                         recv_voucher='R2', receiver='D', recv_cf='CF4', recv_amount=-200,
                         match_type='宽松匹配'),
        ],
        unmatched_pay=[],
        unmatched_recv=[],
    )

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        temp_path = f.name
    try:
        exporter.export(result, temp_path)
        wb = openpyxl.load_workbook(temp_path)
        ws = wb['汇总表']
        # 匹配成功数 = 2
        assert ws.cell(4, 2).value == 2
        # 匹配成功率 = 100%
        assert ws.cell(7, 2).value == '100.0%'
        # 付款总额 = 300
        assert ws.cell(10, 2).value == 300
        # 收款总额 = -300
        assert ws.cell(11, 2).value == -300
        # 差额 = 0
        assert ws.cell(12, 2).value == 0
        wb.close()
    finally:
        os.unlink(temp_path)
