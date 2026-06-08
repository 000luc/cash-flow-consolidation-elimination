# cashflow_tool/exporter.py
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from cashflow_tool.models import MatchResult


class ExcelExporter:
    """生成现金流内部抵消结果 Excel

    3个 sheet:
    1. 现金流内部抵消明细 - 匹配成功的行
    2. 未匹配明细 - 未匹配的记录
    3. 汇总表 - 统计数据
    """

    HEADER_FONT = Font(name='微软雅黑', bold=True, size=11)
    DATA_FONT = Font(name='微软雅黑', size=10)
    BOLD_FONT = Font(name='微软雅黑', bold=True, size=10)
    THIN_BORDER = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )
    HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    SUMMARY_LABEL_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    CNTR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

    def export(self, result: MatchResult, output_path: str):
        """生成完整的输出 Excel"""
        wb = openpyxl.Workbook()
        self._write_matched(wb, result.matched)
        self._write_unmatched(wb, result.unmatched_pay, result.unmatched_recv)
        self._write_summary(wb, result)
        wb.save(output_path)

    def _write_matched(self, wb, matched):
        """Sheet1: 抵消明细"""
        ws = wb.active
        ws.title = '现金流内部抵消明细'

        headers = ['凭证号', '交易公司名称-付款方', '现金流指标-付款', '内部付款额',
                   '凭证号', '对方交易公司名称-收款方', '现金流指标-收款', '内部收款额']
        self._write_header(ws, headers)

        total_pay = total_recv = 0.0
        for i, m in enumerate(matched):
            r = i + 2
            total_pay += m.pay_amount
            total_recv += m.recv_amount
            vals = [m.pay_voucher or '', m.payer, m.pay_cf, round(m.pay_amount, 2),
                    m.recv_voucher or '', m.receiver, m.recv_cf, round(m.recv_amount, 2)]
            self._write_row(ws, r, vals, [4, 8])

        # 合计行
        tr = len(matched) + 2
        self._write_row(ws, tr, ['合计', '', '', round(total_pay, 2),
                                  '', '', '', round(total_recv, 2)], [4, 8], is_total=True)

        self._set_column_widths(ws, [20, 36, 55, 18, 20, 36, 55, 18])
        ws.freeze_panes = 'A2'

    def _write_unmatched(self, wb, unp, unr):
        """Sheet2: 未匹配明细"""
        ws = wb.create_sheet('未匹配明细')
        headers = ['公司名称', '对方公司', '现金流指标', '金额', '凭证号', '方向', '数据来源']
        self._write_header(ws, headers)

        r = 2
        for rec in unp + unr:
            vals = [rec.company, rec.counterparty, rec.cf_full,
                    round(rec.amount, 2), rec.voucher_no, rec.direction, rec.source]
            self._write_row(ws, r, vals, [4])
            r += 1

        if r == 2:
            ws.cell(row=2, column=1, value='（无未匹配记录）').font = self.DATA_FONT

        self._set_column_widths(ws, [36, 36, 55, 18, 22, 10, 10])
        ws.freeze_panes = 'A2'

    def _write_summary(self, wb, result: MatchResult):
        """Sheet3: 汇总表"""
        ws = wb.create_sheet('汇总表')

        title_font = Font(name='微软雅黑', bold=True, size=14)
        section_font = Font(name='微软雅黑', bold=True, size=11)

        ws.cell(1, 1, '现金流内部抵消汇总表').font = title_font
        ws.merge_cells('A1:D1')

        # 基本统计
        ws.cell(3, 1, '一、基本统计').font = section_font
        stats = [
            ('匹配成功（对）', len(result.matched)),
            ('未匹配付款（条）', len(result.unmatched_pay)),
            ('未匹配收款（条）', len(result.unmatched_recv)),
            ('匹配成功率', f'{self._calc_rate(result):.1f}%'),
        ]
        for i, (label, value) in enumerate(stats):
            ws.cell(4 + i, 1, label).font = self.BOLD_FONT
            ws.cell(4 + i, 1).fill = self.SUMMARY_LABEL_FILL
            ws.cell(4 + i, 1).border = self.THIN_BORDER
            ws.cell(4 + i, 2, value).font = self.DATA_FONT
            ws.cell(4 + i, 2).border = self.THIN_BORDER

        # 金额汇总
        total_pay = round(sum(m.pay_amount for m in result.matched), 2)
        total_recv = round(sum(m.recv_amount for m in result.matched), 2)
        ws.cell(9, 1, '二、金额汇总').font = section_font
        amt_stats = [
            ('内部付款总额', total_pay),
            ('内部收款总额', total_recv),
            ('差额', round(total_pay + total_recv, 2)),
        ]
        for i, (label, value) in enumerate(amt_stats):
            r = 10 + i
            ws.cell(r, 1, label).font = self.BOLD_FONT
            ws.cell(r, 1).fill = self.SUMMARY_LABEL_FILL
            ws.cell(r, 1).border = self.THIN_BORDER
            ws.cell(r, 2, value).font = self.DATA_FONT
            ws.cell(r, 2).number_format = '#,##0.00'
            ws.cell(r, 2).border = self.THIN_BORDER

        self._set_column_widths(ws, [30, 20])

    def _calc_rate(self, result: MatchResult) -> float:
        total = len(result.matched) + len(result.unmatched_pay) + len(result.unmatched_recv)
        if total == 0:
            return 100.0
        return len(result.matched) / total * 100

    def _write_header(self, ws, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = self.CNTR_ALIGN
            cell.fill = self.HEADER_FILL

    def _write_row(self, ws, row_num: int, vals: list,
                   num_cols: list[int] | None = None, is_total: bool = False):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=c, value=v if v != '' else None)
            cell.font = self.BOLD_FONT if is_total else self.DATA_FONT
            cell.border = self.THIN_BORDER
            if c in (2, 3, 6, 7):
                cell.alignment = self.LEFT_ALIGN
            elif num_cols and c in num_cols:
                cell.alignment = self.RIGHT_ALIGN
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = self.CNTR_ALIGN

    def _set_column_widths(self, ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
